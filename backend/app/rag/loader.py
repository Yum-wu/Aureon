"""
Document loader for RAG system.
Loads Markdown blog posts, parses frontmatter, and splits into chunks.
"""

import re
from typing import List, Dict, Any
from pathlib import Path

import structlog

from app.utils.lang_detect import detect_language as _detect_text_language

logger = structlog.get_logger()


def detect_doc_language(content: str, frontmatter_lang: str = None) -> str:
    """Detect document language, preferring frontmatter over auto-detection.

    Uses the shared lang_detect utility for consistent detection logic.
    """
    if frontmatter_lang in ("zh", "en"):
        return frontmatter_lang
    return _detect_text_language(content[:500])


def parse_frontmatter(content: str) -> tuple[Dict[str, Any], str]:
    """Parse YAML frontmatter from Markdown content. Return (metadata, body)."""
    match = re.match(r"^---\s*\n(.*?)\n---\s*\n(.*)", content, re.DOTALL)
    if not match:
        return {}, content

    frontmatter_text = match.group(1)
    body = match.group(2).strip()

    metadata = {}
    for line in frontmatter_text.strip().split("\n"):
        if ":" in line:
            key, _, value = line.partition(":")
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            # Parse lists like tags: [AI, Hermes Agent]
            if value.startswith("[") and value.endswith("]"):
                value = [v.strip().strip('"').strip("'") for v in value[1:-1].split(",")]
            metadata[key] = value

    return metadata, body


def load_single_document(filepath: str) -> Dict[str, Any]:
    """Load a single document and return {metadata, content}.

    Supports: .md, .txt, .pdf, .docx, .xlsx/.xls
    Args:
        filepath: Absolute path to file.
    Returns:
        dict with keys: metadata, content
    """
    fpath = Path(filepath)
    suffix = fpath.suffix.lower()

    if suffix == ".md":
        content = fpath.read_text(encoding="utf-8")
        metadata, body = parse_frontmatter(content)
        lang = detect_doc_language(body, metadata.get("lang"))
        return {
            "metadata": {
                "source": fpath.name,
                "title": metadata.get("title", fpath.stem),
                "slug": metadata.get("slug", fpath.stem),
                "tags": metadata.get("tags", []),
                "category": metadata.get("category", ""),
                "filepath": str(fpath),
                "language": lang,
                "uploaded": True,
            },
            "content": body,
        }
    elif suffix == ".txt":
        content = fpath.read_text(encoding="utf-8")
        return {
            "metadata": {
                "source": fpath.name,
                "title": fpath.stem,
                "slug": fpath.stem,
                "tags": [],
                "category": "upload",
                "filepath": str(fpath),
                "language": _detect_text_language(content[:500]),
                "uploaded": True,
            },
            "content": content,
        }
    elif suffix == ".pdf":
        result = load_pdf(filepath)
        return {
            "metadata": {**result["metadata"], "slug": fpath.stem, "tags": [], "category": "upload", "filepath": str(fpath), "uploaded": True},
            "content": result["content"],
        }
    elif suffix == ".docx":
        result = load_docx(filepath)
        return {
            "metadata": {**result["metadata"], "slug": fpath.stem, "tags": [], "category": "upload", "filepath": str(fpath), "uploaded": True},
            "content": result["content"],
        }
    elif suffix in (".xlsx", ".xls"):
        result = load_excel(filepath)
        return {
            "metadata": {**result["metadata"], "slug": fpath.stem, "tags": [], "category": "upload", "filepath": str(fpath), "uploaded": True},
            "content": result["content"],
        }
    else:
        raise ValueError(f"Unsupported file type: {suffix}")


def load_pdf(filepath: str) -> Dict[str, Any]:
    """Load a PDF file, extract text from all pages.

    Returns dict with keys: content, metadata
    """
    from pypdf import PdfReader
    fpath = Path(filepath)
    reader = PdfReader(str(fpath))
    pages = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            pages.append(text.strip())
    content = "\n\n".join(pages)
    lang = _detect_text_language(content[:500]) if content else "en"
    return {
        "content": content,
        "metadata": {
            "source": fpath.name,
            "title": fpath.stem,
            "language": lang,
            "file_type": "pdf",
        },
    }


def load_docx(filepath: str) -> Dict[str, Any]:
    """Load a .docx file, extract paragraph text.

    Returns dict with keys: content, metadata
    """
    from docx import Document
    fpath = Path(filepath)
    doc = Document(str(fpath))
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    content = "\n\n".join(paragraphs)
    lang = _detect_text_language(content[:500]) if content else "en"
    return {
        "content": content,
        "metadata": {
            "source": fpath.name,
            "title": fpath.stem,
            "language": lang,
            "file_type": "docx",
        },
    }


def load_excel(filepath: str) -> Dict[str, Any]:
    """Load an Excel file, convert rows to 'col: value' text lines.

    Returns dict with keys: content, metadata
    """
    from openpyxl import load_workbook
    fpath = Path(filepath)
    wb = load_workbook(str(fpath), read_only=True, data_only=True)
    lines = []
    for ws in wb.worksheets:
        headers = []
        for row in ws.iter_rows(values_only=True):
            if not headers:
                headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(row)]
                continue
            parts = []
            for h, v in zip(headers, row):
                if v is not None:
                    parts.append(f"{h}: {v}")
            if parts:
                lines.append(", ".join(parts))
    content = "\n".join(lines)
    lang = _detect_text_language(content[:500]) if content else "en"
    wb.close()
    return {
        "content": content,
        "metadata": {
            "source": fpath.name,
            "title": fpath.stem,
            "language": lang,
            "file_type": "xlsx",
        },
    }


def load_markdown_files(articles_dir: str) -> List[Dict[str, Any]]:
    """Load all Markdown files from directory. Return list of {metadata, content, filepath}."""
    docs = []
    path = Path(articles_dir)
    if not path.exists():
        logger.warning("Articles dir not found: %s", articles_dir)
        return docs

    for fpath in sorted(path.rglob("*.md")):
        content = fpath.read_text(encoding="utf-8")
        metadata, body = parse_frontmatter(content)
        lang = detect_doc_language(body, metadata.get("lang"))
        doc = {
            "metadata": {
                "source": fpath.name,
                "title": metadata.get("title", fpath.stem),
                "slug": metadata.get("slug", fpath.stem),
                "tags": metadata.get("tags", []),
                "category": metadata.get("category", ""),
                "filepath": str(fpath),
                "language": lang,
            },
            "content": body,
        }
        docs.append(doc)

    logger.info("Loaded %d documents from %s", len(docs), articles_dir)
    return docs
